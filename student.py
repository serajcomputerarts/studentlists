# for using this program you have to install some files
# use pip install openpyxl 
# do the same for other files
# after installing use :
# python student.py test
from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk
import sys
from persiantools.jdatetime import JalaliDate
from datetime import datetime
from functools import partial
# this is main class 
class StudentProject:
    exfilename=sys.argv[1]+".xlsx"
    # we will use these two files for reading excel file
    # multiple methods will use it so I put them here
    wb = load_workbook(filename = exfilename)
    ws=wb.active
    # ls will present the all students in excel file
    ls=list()
    # mylist is the list which we will save to new coulumn
    mylist=list()
    # this is a counter for our list 
    i=0
    # here are some variables for user interface
    root=Tk()
    mylabel=ttk.Label()
    # this is empty label which will be use in grid
    blabel=ttk.Label()
    def __init__(self):
        # for start lets get student names 
        self.ls=self.getstudentnames()
        # and persian current date
        self.mylist.append(self.getdate())
       
    # this methid will get all student names from excel 
    # this method returns a list named ls 
    def getstudentnames(self):
        ls=list()
        rownum=self.ws.max_row
        for i in range(2,rownum+1):
          ls.append(self.ws.cell(i,2).value)
        return ls
    # this method showes gui 
    # I tried to write some code but needs improve     
    def showgui(self):
        
        window_height = 250
        window_width = 600
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x_cordinate = int((screen_width/2) - (window_width/2))
        y_cordinate = int((screen_height/2) - (window_height/2))
        self.root.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
        photo1 = PhotoImage(file = "pics/tik.png")
        photo2 = PhotoImage(file = "pics/utik.png")
        photo3 = PhotoImage(file = "pics/dolar.png")
        photo4 = PhotoImage(file = "pics/bad.png")
        self.mylabel.config(text=self.ls[self.i],font=("B Yekan",25))
        self.blabel.config(text=" ",font=("B Yekan",40))
        self.blabel.grid(column=1, row=1)
        self.blabel.grid(column=1, row=2)
        self.blabel.grid(column=1, row=3)
        self.mylabel.grid(column=5, row=4)
        #self.s(self.root, text=self.ls[self.i],font=("B Yekan",25)).grid(column=2, row=3)
        ttk.Button(self.root, text="Quit", image=photo1,command=partial(self.setmark,0.25)).grid(column=1, row=4)
        ttk.Button(self.root, text="Quit", image=photo2,command=partial(self.setmark,0)).grid(column=2, row=4)
        ttk.Button(self.root, text="Quit", image=photo3,command=partial(self.setmark,0.5)).grid(column=3, row=4)
        ttk.Button(self.root, text="Quit", image=photo4,command=partial(self.setmark,-0.25)).grid(column=4, row=4)
        self.root.mainloop()
    # this method will add mark to mylist and call next student 
    def setmark(self,mark):
        self.mylist.append(mark)
        self.nextstudent()
                
    #this method will show next student
    def nextstudent(self):
        
        if (self.i<len(self.ls)-1):
                self.i+=1
                self.mylabel.config(text=self.ls[self.i],font=("B Yekan",25))
                self.mylabel.grid(column=5, row=4)
        else:    
            self.mylabel.config(text="به امید دیدار در جلسه بعد",font=("B Yekan",25))
            self.mylabel.grid(column=5, row=4)
            self.writetoexcel()
            self.root.quit()
    # this method writes mylist to excel file
    def writetoexcel(self):
        clm=self.ws.max_column+1
        i=1
        for x in range(len(self.mylist)):
            t=self.ws.cell(row=i,column=clm)
            t.value=self.mylist[x]
            i=i+1
        self.wb.save(self.exfilename)
    # this method gets systemdate and convert is to persian
    def getdate(self):
        today=datetime.today()
        return str(JalaliDate(today))
# ----- class finished lets test
ob=StudentProject()
ob.showgui()








